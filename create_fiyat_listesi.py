import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
import os

# ── Kurumsal Renkler ──────────────────────────────────────────────
NAVY      = "0D1F3C"   # Ana lacivert
NAVY2     = "142444"
NAVY3     = "1A2E55"
GREEN     = "4A9E6B"   # Aksan yeşil
GREEN2    = "3A8A5A"
LIGHT     = "F5F7FA"   # Açık arka plan
BORDER_C  = "E2E8F0"
GRAY      = "64748B"
WHITE     = "FFFFFF"
YELLOW_HL = "FFF176"   # Vurgulu satırlar için (listede sarı işaretli)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Fiyat Listesi"

# ── Sayfa ayarları ────────────────────────────────────────────────
ws.page_setup.orientation = "portrait"
ws.page_setup.paperSize   = ws.PAPERSIZE_A4
ws.print_area             = "A1:E100"
ws.page_margins.left      = 0.5
ws.page_margins.right     = 0.5
ws.page_margins.top       = 0.7
ws.page_margins.bottom    = 0.7

# ── Sütun genişlikleri ────────────────────────────────────────────
ws.column_dimensions["A"].width = 14    # Ürün kodu
ws.column_dimensions["B"].width = 52    # Ürün adı
ws.column_dimensions["C"].width = 10    # Adet (çift)
ws.column_dimensions["D"].width = 12    # Birim fiyat
ws.column_dimensions["E"].width = 14    # Notlar / extra

# ── Yardımcı fonksiyonlar ─────────────────────────────────────────
def navy_fill(hex_color=NAVY):
    return PatternFill("solid", fgColor=hex_color)

def thin_border(color=BORDER_C):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def bottom_border(color=BORDER_C):
    s = Side(style="thin", color=color)
    return Border(bottom=s)

def cell(row, col, value="", bold=False, size=10, color=NAVY,
         bg=None, align="left", italic=False, wrap=False, border=None, num_format=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    if border:
        c.border = border
    if num_format:
        c.number_format = num_format
    return c

# ── Logo ──────────────────────────────────────────────────────────
logo_path = os.path.join(os.path.dirname(__file__), "cansular-logo-200x100.png")
if os.path.exists(logo_path):
    # Logo için 5 satır boşluk bırak
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 20
    img = XLImage(logo_path)
    img.width  = 160
    img.height = 80
    ws.add_image(img, "A1")
    START_ROW = 1
else:
    START_ROW = 1

# ── Header bloğu (satır 1-5) ──────────────────────────────────────
# Tüm header hücrelerini lacivert yap
for r in range(1, 6):
    ws.row_dimensions[r].height = 18
    for c_idx in range(1, 6):
        cell(r, c_idx, bg=NAVY)

# Sağ taraf: şirket bilgileri (logo yoksa sol tarafa da yazılır)
ws.merge_cells("C1:E1")
ws.merge_cells("C2:E2")
ws.merge_cells("C3:E3")
ws.merge_cells("C4:E4")
ws.merge_cells("C5:E5")

cell(1, 3, "CANSULAR HIRDAVAT SANAYİ A.Ş.", bold=True, size=13,
     color=WHITE, bg=NAVY, align="right")
cell(2, 3, "Koruyucu Eldiven Uzmanı", italic=True, size=9,
     color="4A9E6B", bg=NAVY, align="right")
cell(3, 3, "www.cansular.com", size=9, color="94A3B8", bg=NAVY, align="right")
cell(4, 3, "", bg=NAVY)
cell(5, 3, "Mart 2026 – Geçerli Fiyat Listesi", bold=True, size=9,
     color="94A3B8", bg=NAVY, align="right")

# ── Başlık bandı (satır 6) ────────────────────────────────────────
ws.merge_cells("A6:E6")
ws.row_dimensions[6].height = 36
c6 = ws.cell(row=6, column=1, value="FİYAT LİSTESİ / PRICE LIST")
c6.font      = Font(name="Calibri", bold=True, size=16, color=WHITE)
c6.fill      = PatternFill("solid", fgColor=GREEN)
c6.alignment = Alignment(horizontal="center", vertical="center")

# ── Tablo başlıkları (satır 7) ────────────────────────────────────
ws.row_dimensions[7].height = 22
headers = ["ÜRÜN KODU", "ÜRÜN ADI", "KUTU ADEDİ", "BİRİM FİYAT (₺)", "NOTLAR"]
for ci, h in enumerate(headers, 1):
    cell(7, ci, h, bold=True, size=9, color=WHITE, bg=NAVY2,
         align="center", border=thin_border(NAVY3))

# ── Ürün verileri ─────────────────────────────────────────────────
# (kod, ad, adet, fiyat, notlar, kategori_başlık, vurgulu)
# kategori_başlık=True → kategori satırı
# vurgulu=True          → sarı vurgulu satır

URUNLER = [
    # --- Kategori 1 ---
    (None, "NİTRİL ELDİVEN", None, None, None, True, False),
    ("C-55",  "EKSTRA UNIVERSAL NİTRİL ELDİVEN",              "240 ÇİFT", 50.95, "",  False, False),
    ("C-25",  "TECHNO UNIVERSAL NİTRİL TAMKAPLI ELDİVEN",     "288 ÇİFT", 42.95, "",  False, False),
    ("C-20",  "TECHNO UNIVERSAL NİTRİL ELDİVEN",              "288 ÇİFT", 39.75, "",  False, False),
    ("C-15",  "STANDART UNIVERSAL NİTRİL ELDİVEN",            "288 ÇİFT", 37.20, "",  False, False),

    # --- Kategori 2 ---
    (None, "MONTAJ & PU ELDİVEN", None, None, None, True, False),
    ("C-75",     "POWER UNIVERSAL NİTRİL ELDİVEN (Petrolcü)",                   "144 ÇİFT", 69.85, "",  False, False),
    ("8701-PU/A","UNIVERSAL TECHNO ELDİVEN",                                     "300 ÇİFT", 25.85, "",  False, False),
    ("8802",     "MONTAJ UNIVERSAL ELDİVEN (SİYAH, GRİ, SARI, KIRMIZI, MAVİ)", "300 ÇİFT", 35.65, "",  False, False),
    ("8802-B",   "MONTAJ UNIVERSAL ELDİVEN (TURUNCU, MAVİ)",                    "300 ÇİFT", 38.95, "",  False, False),
    ("8804",     "TAMKAPLI UNIVERSAL NİTRİL ELDİVEN",                           "240 ÇİFT", 49.95, "",  False, False),

    # --- Kategori 3 ---
    (None, "KİRÇİLLİ ELDİVEN", None, None, None, True, False),
    ("6121-RL",  "KİRÇİLLİ UNIVERSAL NİTRİL ELDİVEN",        "144 ÇİFT", 51.95, "",  False, False),

    # --- Kategori 4 ---
    (None, "KESİLMEZ ELDİVEN", None, None, None, True, False),
    ("8701-PU-3","KESİLMEZ UNIVERSAL ELDİVEN",                "—",        None,  "Fiyat için iletişime geçin", False, False),

    # --- Kategori 5 (vurgulu) ---
    (None, "LATEKS KAPLAMA ELDİVEN", None, None, None, True, False),
    ("6101",     "LATEKS KAPLAMA UNİVERSAL ELDİVEN YESİL",   "180 ÇİFT", 50.95, "",  False, True),
    ("6155",     "LATEKS KAPLAMA UNİVERSAL ELDİVEN KİRÇİLLİ","144 ÇİFT", 45.35, "",  False, True),
    ("6600",     "LATEKS KAPLAMA UNIVERSAL ELDİVEN 3/4 KAPLI","144 ÇİFT", None,  "Fiyat için iletişime geçin", False, True),
    ("8906-DC",  "UNIVERSAL ÇİFT KAPLI MAVİ ELDİVEN",        "144 ÇİFT", 69.85, "",  False, True),
    ("C-40",     "GARDEN UNIVERSAL NİTRİL ELDİVEN",           "288 ÇİFT", 28.85, "",  False, True),
    ("C-432",    "CAMCI UNIVERSAL TECHNO ELDİVEN (KIRMIZI, GRİ)", "288 ÇİFT", 38.95, "", False, True),
]

current_row = 8
alt = False  # zebra striping

for (kod, ad, adet, fiyat, notlar, is_cat, is_hl) in URUNLER:
    ws.row_dimensions[current_row].height = 20 if is_cat else 18

    if is_cat:
        # Kategori başlık satırı
        ws.merge_cells(f"A{current_row}:E{current_row}")
        c = ws.cell(row=current_row, column=1, value=f"  {ad}")
        c.font      = Font(name="Calibri", bold=True, size=9, color=WHITE)
        c.fill      = PatternFill("solid", fgColor=NAVY3)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border    = thin_border(NAVY)
        alt = False  # zebra sıfırla
    else:
        if is_hl:
            row_bg = "FFFDE7"   # hafif sarı
        elif alt:
            row_bg = LIGHT
        else:
            row_bg = WHITE
        alt = not alt

        brd = thin_border(BORDER_C)

        cell(current_row, 1, kod or "",  size=9, color=NAVY, bg=row_bg, align="center", border=brd, bold=True)
        cell(current_row, 2, ad or "",   size=9, color=NAVY, bg=row_bg, align="left",   border=brd)
        cell(current_row, 3, adet or "", size=9, color=GRAY, bg=row_bg, align="center", border=brd)

        if fiyat is not None:
            fc = ws.cell(row=current_row, column=4, value=fiyat)
            fc.font         = Font(name="Calibri", size=9, bold=True, color=GREEN2)
            fc.fill         = PatternFill("solid", fgColor=row_bg)
            fc.alignment    = Alignment(horizontal="right", vertical="center")
            fc.number_format = '#,##0.00 ₺'
            fc.border        = brd
        else:
            cell(current_row, 4, "—", size=9, color=GRAY, bg=row_bg, align="center", border=brd)

        cell(current_row, 5, notlar or "", size=8, color=GRAY, bg=row_bg,
             align="left", border=brd, italic=True)

    current_row += 1

# ── Alt bilgi satırı ─────────────────────────────────────────────
current_row += 1
ws.merge_cells(f"A{current_row}:E{current_row}")
ws.row_dimensions[current_row].height = 30
fc = ws.cell(row=current_row, column=1,
             value="Fiyatlar KDV hariçtir. Tüm fiyatlar Türk Lirası (₺) cinsindendir. • Tel: +90 212 xxx xx xx  •  info@cansular.com")
fc.font      = Font(name="Calibri", size=8, italic=True, color=WHITE)
fc.fill      = PatternFill("solid", fgColor=NAVY)
fc.alignment = Alignment(horizontal="center", vertical="center")

# ── Kaydet ────────────────────────────────────────────────────────
out_path = os.path.join(os.path.dirname(__file__), "Cansular_Fiyat_Listesi_Mart2026.xlsx")
wb.save(out_path)
print(f"Kaydedildi: {out_path}")
